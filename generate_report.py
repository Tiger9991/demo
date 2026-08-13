import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "تقرير التعديلات"

# Set RTL reading order for the sheet
ws.sheet_view.rightToLeft = True

# --- Color Palette ---
HEADER_BG     = "1F3864"   # dark navy
HEADER_FG     = "FFFFFF"
SUB_HDR_BG    = "2E75B6"   # medium blue
SUB_HDR_FG    = "FFFFFF"
ROW_ODD       = "EBF3FB"   # light blue
ROW_EVEN      = "FFFFFF"
TITLE_BG      = "0D47A1"   # deep blue
TITLE_FG      = "FFFFFF"
BORDER_COLOR  = "2E75B6"

# --- Helper: thin border ---
def make_border(color=BORDER_COLOR):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

# --- Helper: fill ---
def solid_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ================================================================
# Row 1: Main Title
# ================================================================
ws.merge_cells("A1:C1")
title_cell = ws["A1"]
title_cell.value = "تقرير التعديلات على النظام"
title_cell.font = Font(name="Cairo", bold=True, size=18, color=TITLE_FG)
title_cell.fill = solid_fill(TITLE_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
title_cell.border = make_border("FFFFFF")
ws.row_dimensions[1].height = 45

# ================================================================
# Row 2: Date
# ================================================================
ws.merge_cells("A2:C2")
date_cell = ws["A2"]
date_cell.value = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d')}"
date_cell.font = Font(name="Cairo", bold=False, size=11, color="444444")
date_cell.fill = solid_fill("D6E4F7")
date_cell.alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)
date_cell.border = make_border("B0C4DE")
ws.row_dimensions[2].height = 25

# ================================================================
# Row 3: blank spacer
# ================================================================
ws.row_dimensions[3].height = 8

# ================================================================
# Row 4: Column Headers
# ================================================================
headers = ["العنصر", "سابقاً", "ما تم تعديله"]
header_row = 4
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col_idx, value=header)
    cell.font = Font(name="Cairo", bold=True, size=13, color=HEADER_FG)
    cell.fill = solid_fill(HEADER_BG)
    cell.alignment = Alignment(
        horizontal="center", vertical="center",
        wrap_text=True, readingOrder=2
    )
    cell.border = make_border("FFFFFF")
ws.row_dimensions[header_row].height = 35

# ================================================================
# Data Rows
# ================================================================
data = [
    (
        "المحطات المتصلة",
        "تحتوي على وحدة قياس قوة الإشارة بالإضافة إلى التعبير عن قوة الإشارة باللغة الإنجليزية",
        "تم حذف قياس وحدة الإشارة بالإضافة إلى تغيير التعبير إلى اللغة العربية",
    ),
    (
        "المحطات المتصلة\n(آخر دخول)",
        "آخر دخول بتوقيت السيرفر",
        "تم التعديل إلى التوقيت المصري",
    ),
    (
        "إجمالي نشاط القوارض",
        "توقيت آخر زيارة بتوقيت السيرفر",
        "تم التعديل إلى التوقيت المصري",
    ),
    (
        "محطات إعادة تعبئة الطعوم",
        "التوقيت بتوقيت السيرفر",
        "تم التعديل إلى التوقيت المصري",
    ),
    (
        "آخر توقيت تنبيهات",
        "التوقيت بتوقيت السيرفر وتواجد عمود قوة الإشارة",
        "تم التعديل إلى التوقيت المصري\nتم حذف عمود قوة الإشارة",
    ),
    (
        "مؤشر شدة الإصابة",
        "ألوانه تماثل الألوان الظاهرة على الخريطة",
        "تم التعديل إلى ألوان مختلفة",
    ),
    (
        "مخطط الإصابة الزمني",
        "التوقيت بتوقيت السيرفر",
        "تم التعديل إلى التوقيت المصري",
    ),
    (
        "القائمة",
        "تظهر بكافة عناصرها",
        "يظهر فقط العنصر الأول",
    ),
]

fills = [solid_fill(ROW_ODD), solid_fill(ROW_EVEN)]

for i, (element, before, after) in enumerate(data):
    row_num = header_row + 1 + i
    row_fill = fills[i % 2]

    values = [element, before, after]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = Font(name="Cairo", size=11, color="1A1A2E")
        cell.fill = row_fill
        cell.alignment = Alignment(
            horizontal="right", vertical="center",
            wrap_text=True, readingOrder=2,
            indent=1
        )
        cell.border = make_border()

    ws.row_dimensions[row_num].height = 55

# ================================================================
# Column Widths (A=العنصر, B=سابقاً, C=ما تم تعديله)
# ================================================================
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 50

# ================================================================
# Freeze panes below header row
# ================================================================
ws.freeze_panes = "A5"

# ================================================================
# Save
# ================================================================
output_path = r"d:\system\تقرير_التعديلات.xlsx"
wb.save(output_path)
import sys
sys.stdout.buffer.write(("Saved: " + output_path + "\n").encode("utf-8"))
